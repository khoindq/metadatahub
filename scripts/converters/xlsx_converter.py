"""Excel → JSON converter using openpyxl.

Extracts sheet names, headers, sample rows, column types, and computed stats.
Produces a structured JSON per sheet for indexing with rich schema metadata.
"""

from datetime import datetime, date
from pathlib import Path
from typing import Optional, Any

from openpyxl import load_workbook


def convert(filepath: Path, output_dir: Optional[Path] = None) -> dict:
    """Convert an Excel file to structured JSON with rich metadata.

    Args:
        filepath: Path to the .xlsx file.
        output_dir: Directory to write converted JSON files.

    Returns:
        dict with keys:
            sheets: list of sheet info dicts (with columns, stats, sample_data)
            output_files: list of written file paths
    """
    import json

    filepath = Path(filepath)
    wb = load_workbook(filepath, read_only=True, data_only=True)

    sheets = []
    output_files = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            sheets.append({
                "name": sheet_name,
                "headers": [],
                "row_count": 0,
                "sample_rows": [],
                "column_count": 0,
                "columns": [],
                "stats": {},
                "sample_data": "",
            })
            continue

        # First row as headers
        headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
        data_rows = rows[1:]

        # Analyze columns - types and stats
        columns = _analyze_columns(headers, data_rows)

        # Sample rows (first 5)
        sample_rows = []
        for row in data_rows[:5]:
            sample_row = {}
            for header, val in zip(headers, row):
                sample_row[header] = _serialize_value(val)
            sample_rows.append(sample_row)

        # Compute sheet-level stats
        stats = _compute_sheet_stats(columns, len(data_rows))

        # Format sample_data string (PageIndex-style)
        sample_data = _format_sample_data(headers, sample_rows)

        sheet_info = {
            "name": sheet_name,
            "headers": headers,
            "row_count": len(data_rows),
            "column_count": len(headers),
            "columns": columns,
            "sample_rows": sample_rows,
            "sample_data": sample_data,
            "stats": stats,
        }
        sheets.append(sheet_info)

    wb.close()

    result = {
        "sheets": sheets,
        "sheet_count": len(sheets),
        "output_files": [],
    }

    if output_dir:
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        for sheet_info in sheets:
            safe_name = sheet_info["name"].replace("/", "_").replace(" ", "_").lower()
            out_path = output_dir / f"sheet_{safe_name}.json"
            out_path.write_text(json.dumps(sheet_info, indent=2, default=str) + "\n")
            result["output_files"].append(str(out_path))

    return result


def _analyze_columns(headers: list[str], data_rows: list[tuple]) -> list[dict]:
    """Analyze each column for type, stats, and sample values."""
    columns = []

    for col_idx, header in enumerate(headers):
        col_values = [row[col_idx] if col_idx < len(row) else None for row in data_rows]
        non_null_values = [v for v in col_values if v is not None]

        col_info = {
            "name": header,
            "type": "unknown",
            "non_null_count": len(non_null_values),
            "null_count": len(col_values) - len(non_null_values),
        }

        if not non_null_values:
            col_info["type"] = "empty"
            columns.append(col_info)
            continue

        # Detect column type
        col_type, typed_values = _detect_column_type(non_null_values)
        col_info["type"] = col_type

        # Compute type-specific stats
        if col_type == "numeric":
            col_info["stats"] = _compute_numeric_stats(typed_values)
        elif col_type == "date":
            col_info["stats"] = _compute_date_stats(typed_values)
        elif col_type == "text":
            col_info["stats"] = _compute_text_stats(typed_values)

        # Unique values for categorical detection
        unique_vals = set(str(v) for v in non_null_values[:1000])  # limit for perf
        col_info["unique_count"] = len(unique_vals)

        # Mark as categorical if few unique values relative to total
        if col_type == "text" and len(unique_vals) <= 20 and len(non_null_values) > 10:
            col_info["is_categorical"] = True
            col_info["categories"] = sorted(unique_vals)[:20]

        columns.append(col_info)

    return columns


def _detect_column_type(values: list) -> tuple[str, list]:
    """Detect the dominant type of a column and return typed values."""
    numeric_count = 0
    date_count = 0
    text_count = 0

    numeric_vals = []
    date_vals = []

    for v in values:
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            numeric_count += 1
            numeric_vals.append(float(v))
        elif isinstance(v, (datetime, date)):
            date_count += 1
            date_vals.append(v)
        else:
            # Try parsing as number
            try:
                numeric_vals.append(float(v))
                numeric_count += 1
            except (ValueError, TypeError):
                text_count += 1

    total = len(values)
    # Require >70% consistency for type detection
    if numeric_count / total > 0.7:
        return "numeric", numeric_vals
    elif date_count / total > 0.7:
        return "date", date_vals
    else:
        return "text", values


def _compute_numeric_stats(values: list[float]) -> dict:
    """Compute statistics for numeric columns."""
    if not values:
        return {}

    sorted_vals = sorted(values)
    n = len(values)
    total = sum(values)

    stats = {
        "min": sorted_vals[0],
        "max": sorted_vals[-1],
        "sum": total,
        "avg": total / n,
        "count": n,
    }

    # Median
    if n % 2 == 0:
        stats["median"] = (sorted_vals[n // 2 - 1] + sorted_vals[n // 2]) / 2
    else:
        stats["median"] = sorted_vals[n // 2]

    # Format large numbers nicely
    if stats["sum"] >= 1_000_000:
        stats["sum_formatted"] = _format_number(stats["sum"])
    if stats["avg"] >= 1_000:
        stats["avg_formatted"] = _format_number(stats["avg"])

    return stats


def _compute_date_stats(values: list) -> dict:
    """Compute statistics for date columns."""
    if not values:
        return {}

    # Convert to dates for comparison
    date_vals = []
    for v in values:
        if isinstance(v, datetime):
            date_vals.append(v.date() if hasattr(v, 'date') else v)
        elif isinstance(v, date):
            date_vals.append(v)

    if not date_vals:
        return {}

    sorted_dates = sorted(date_vals)
    return {
        "min_date": sorted_dates[0].isoformat(),
        "max_date": sorted_dates[-1].isoformat(),
        "date_range": f"{sorted_dates[0].isoformat()} to {sorted_dates[-1].isoformat()}",
        "count": len(date_vals),
    }


def _compute_text_stats(values: list) -> dict:
    """Compute statistics for text columns."""
    if not values:
        return {}

    lengths = [len(str(v)) for v in values]
    return {
        "avg_length": sum(lengths) / len(lengths),
        "max_length": max(lengths),
        "count": len(values),
    }


def _compute_sheet_stats(columns: list[dict], row_count: int) -> dict:
    """Compute sheet-level aggregate stats from column analysis."""
    stats = {
        "row_count": row_count,
    }

    # Aggregate numeric columns
    numeric_cols = [c for c in columns if c.get("type") == "numeric" and c.get("stats")]
    if numeric_cols:
        # Find the likely "amount/value" column (largest sum)
        amount_col = max(numeric_cols, key=lambda c: c["stats"].get("sum", 0))
        stats["primary_numeric_column"] = amount_col["name"]
        stats["total_" + amount_col["name"].lower().replace(" ", "_")] = amount_col["stats"]["sum"]
        if "sum_formatted" in amount_col["stats"]:
            stats["total_formatted"] = amount_col["stats"]["sum_formatted"]

    # Aggregate date columns
    date_cols = [c for c in columns if c.get("type") == "date" and c.get("stats")]
    if date_cols:
        all_min = min(c["stats"]["min_date"] for c in date_cols)
        all_max = max(c["stats"]["max_date"] for c in date_cols)
        stats["date_range"] = f"{all_min} to {all_max}"

    # Count categorical columns
    categorical_cols = [c for c in columns if c.get("is_categorical")]
    if categorical_cols:
        stats["categorical_columns"] = [c["name"] for c in categorical_cols]

    return stats


def _format_number(n: float) -> str:
    """Format large numbers with K/M/B suffixes."""
    if abs(n) >= 1_000_000_000:
        return f"${n / 1_000_000_000:.2f}B"
    elif abs(n) >= 1_000_000:
        return f"${n / 1_000_000:.2f}M"
    elif abs(n) >= 1_000:
        return f"${n / 1_000:.1f}K"
    else:
        return f"${n:.2f}"


def _format_sample_data(headers: list[str], sample_rows: list[dict]) -> str:
    """Format sample data as a readable string (PageIndex-style).

    Example: "Row 1: 2025-01-05, Widget Pro, California, $12,500"
    """
    if not sample_rows:
        return ""

    lines = []
    for i, row in enumerate(sample_rows[:3], 1):  # First 3 rows
        vals = []
        for h in headers[:6]:  # First 6 columns
            v = row.get(h)
            if v is not None:
                # Format numbers nicely
                if isinstance(v, float) and v >= 1000:
                    vals.append(_format_number(v))
                elif isinstance(v, float):
                    vals.append(f"{v:.2f}" if v != int(v) else str(int(v)))
                else:
                    vals.append(str(v)[:30])  # Truncate long strings
        lines.append(f"Row {i}: {', '.join(vals)}")

    return "; ".join(lines)


def _serialize_value(val: Any) -> Any:
    """Serialize a cell value to JSON-compatible form."""
    if val is None:
        return None
    if isinstance(val, (int, float, bool)):
        return val
    if isinstance(val, (datetime, date)):
        return val.isoformat()
    return str(val)


def get_sample(filepath: Path, max_rows: int = 5) -> str:
    """Extract a sample from the Excel file for AI sampling.

    Returns sheet names + headers + column types + stats + first N data rows.
    """
    filepath = Path(filepath)
    wb = load_workbook(filepath, read_only=True, data_only=True)

    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            parts.append(f"[Sheet: {sheet_name}]\n(empty)")
            continue

        headers = [str(h) if h is not None else "" for h in rows[0]]
        data_rows = rows[1:]
        header_line = " | ".join(headers)

        parts.append(f"[Sheet: {sheet_name}]")
        parts.append(f"Headers: {header_line}")
        parts.append(f"Total rows: {len(data_rows)}")

        # Add column type hints from first 100 rows
        if data_rows:
            columns = _analyze_columns(headers, data_rows[:100])
            col_types = []
            for col in columns:
                type_str = col["type"]
                if col.get("is_categorical"):
                    cats = col.get("categories", [])[:5]
                    type_str = f"categorical({', '.join(cats)})"
                elif col.get("stats"):
                    if "sum_formatted" in col["stats"]:
                        type_str = f"numeric(total: {col['stats']['sum_formatted']})"
                    elif "date_range" in col["stats"]:
                        type_str = f"date({col['stats']['date_range']})"
                col_types.append(f"{col['name']}: {type_str}")
            parts.append(f"Column types: {'; '.join(col_types)}")

        # Sample rows
        for row in data_rows[:max_rows]:
            vals = []
            for i, v in enumerate(row):
                if v is None:
                    vals.append("")
                elif isinstance(v, (datetime, date)):
                    vals.append(v.isoformat()[:10])
                elif isinstance(v, float) and v >= 1000:
                    vals.append(_format_number(v))
                else:
                    vals.append(str(v)[:30])
            parts.append("  " + " | ".join(vals))

    wb.close()
    return "\n".join(parts)
