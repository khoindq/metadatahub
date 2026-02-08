"""Smart Excel converter using LLM for complex layouts.

For Excel files with non-standard layouts (form-based, mixed content, etc.),
uses Claude to understand and extract structured data.
"""

from pathlib import Path
from typing import Optional
import json

from openpyxl import load_workbook


EXTRACT_SYSTEM_PROMPT = """You are a data extraction expert. Your job is to extract structured information from Excel sheet content.

The user will provide raw cell data from an Excel sheet. You must:
1. Understand the layout (tabular, form-based, or mixed)
2. Extract ALL data into a clean structured format
3. Return valid JSON only

For form-based layouts (key-value pairs scattered in cells), extract as:
{
  "layout": "form",
  "data": {
    "field_name": "value",
    ...
  },
  "sections": [
    {
      "name": "section_name",
      "fields": {"key": "value", ...}
    }
  ]
}

For tabular layouts, extract as:
{
  "layout": "tabular", 
  "columns": ["col1", "col2", ...],
  "row_count": 100,
  "sample_rows": [{"col1": "val", ...}, ...],
  "summary": "description of data"
}

For mixed layouts, combine both approaches.

Always include:
- "summary": 1-2 sentence description
- "key_fields": list of important field names found
"""

EXTRACT_USER_TEMPLATE = """Extract structured data from this Excel sheet.

**Sheet name:** {sheet_name}
**Dimensions:** {rows} rows × {cols} columns

**Raw cell content (row, col → value):**
```
{cell_content}
```

Return the extracted data as JSON only."""


def convert_with_llm(
    filepath: Path,
    client,  # ClaudeClient
    output_dir: Optional[Path] = None,
    max_rows_sample: int = 50,
) -> dict:
    """Convert Excel using LLM to understand and extract data.
    
    Args:
        filepath: Path to Excel file
        client: ClaudeClient for LLM calls
        output_dir: Directory to write output JSON files
        max_rows_sample: Max rows to send to LLM per sheet
        
    Returns:
        dict with sheets data and output files
    """
    filepath = Path(filepath)
    wb = load_workbook(filepath, read_only=True, data_only=True)
    
    sheets = []
    output_files = []
    
    total_sheets = len(wb.sheetnames)
    for idx, sheet_name in enumerate(wb.sheetnames):
        print(f"    [{idx+1}/{total_sheets}] {sheet_name}...", end=" ", flush=True)
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        
        if not rows:
            print("(empty)")
            sheets.append({
                "name": sheet_name,
                "layout": "empty",
                "data": {},
            })
            continue
        
        # Build cell content representation for LLM
        cell_content = _build_cell_content(rows[:max_rows_sample])
        
        # Ask LLM to extract structured data
        extracted = _extract_with_llm(
            client, 
            sheet_name, 
            len(rows), 
            len(rows[0]) if rows else 0,
            cell_content
        )
        
        if extracted:
            print(f"✓ {extracted.get('layout', 'ok')}")
            sheet_info = {
                "name": sheet_name,
                **extracted,
                "total_rows": len(rows),
            }
        else:
            print("✗ fallback")
            # Fallback to raw representation
            sheet_info = {
                "name": sheet_name,
                "layout": "unknown",
                "raw_preview": cell_content[:2000],
                "total_rows": len(rows),
            }
        
        sheets.append(sheet_info)
    
    wb.close()
    
    result = {
        "sheets": sheets,
        "sheet_count": len(sheets),
        "output_files": [],
    }
    
    if output_dir:
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        
        for sheet_info in sheets:
            safe_name = sheet_info["name"].replace("/", "_").replace(" ", "_").lower()
            out_path = output_dir / f"sheet_{safe_name}.json"
            out_path.write_text(
                json.dumps(sheet_info, indent=2, ensure_ascii=False, default=str) + "\n",
                encoding="utf-8"
            )
            result["output_files"].append(str(out_path))
    
    return result


def _build_cell_content(rows: list, max_cols: int = 10) -> str:
    """Build a readable representation of cell content for LLM."""
    lines = []
    for row_idx, row in enumerate(rows):
        for col_idx, val in enumerate(row[:max_cols]):
            if val is not None:
                val_str = str(val)[:100]  # Truncate long values
                lines.append(f"[{row_idx+1},{col_idx+1}] {val_str}")
    return "\n".join(lines)


def _extract_with_llm(
    client,
    sheet_name: str,
    num_rows: int,
    num_cols: int,
    cell_content: str,
) -> Optional[dict]:
    """Use LLM to extract structured data from cell content."""
    prompt = EXTRACT_USER_TEMPLATE.format(
        sheet_name=sheet_name,
        rows=num_rows,
        cols=num_cols,
        cell_content=cell_content,
    )
    
    try:
        result = client.send_json_message(
            prompt=prompt,
            system=EXTRACT_SYSTEM_PROMPT,
            max_tokens=4096,
        )
        
        if result.get("parsed"):
            return result["parsed"]
    except Exception as e:
        print(f"  Warning: LLM extraction failed for {sheet_name}: {e}")
    
    return None


def get_sample(filepath: Path, max_rows: int = 10) -> str:
    """Get a sample for AI analysis (same as regular xlsx_converter)."""
    from scripts.converters.xlsx_converter import get_sample as basic_sample
    return basic_sample(filepath, max_rows)
